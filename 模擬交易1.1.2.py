import fugle_marketdata as fg
import pandas as pd
import yaml
import json
import os
import numpy as np
import openpyxl
import math
import colorama
import subprocess
import sys
import time as time_module
import warnings
from tabulate import tabulate
from openpyxl.styles import PatternFill
from colorama import init, Fore, Style
from datetime import datetime, time, timedelta, date
from fugle_marketdata import RestClient
from fugle_realtime import WebSocketClient
import websocket
import threading
import msvcrt

colorama.init(autoreset=True)
warnings.filterwarnings("ignore", category=FutureWarning)

required_packages = [
    'fugle-marketdata',
    'pandas',
    'pyyaml',
    'colorama',
    'numpy',
    'python-dateutil',
    'tabulate',
    'openpyxl'
]

def install_package(package):
    try:
        subprocess.check_call(
            [sys.executable, "-m", "pip", "install", package],
            stdout=subprocess.DEVNULL,
            stderr=subprocess.DEVNULL
        )
        print(f"{package} 安裝成功")
    except subprocess.CalledProcessError:
        print(f"{package} 安裝失敗")

def check_and_install_packages(packages):
    for package in packages:
        try:
            __import__(package)
            print(f"{package} 已安裝")
        except ImportError:
            install_package(package)

init(autoreset=True)

RED = Fore.RED
GREEN = Fore.GREEN
YELLOW = Fore.YELLOW
BLUE = Fore.BLUE
RESET = Style.RESET_ALL

pd.set_option('future.no_silent_downcasting', True)

def init_fugle_client():
    try:
        config = load_config("config.yaml")
        client = RestClient(api_key=config['api_key'])
        print("=" * 50)
        print("從 config.yaml 載入 API 金鑰")
        print("=" * 50)
        return client, config['api_key']
    except FileNotFoundError:
        print("錯誤：config.yaml 文件不存在。")
        sys.exit(1)
    except KeyError:
        print("錯誤：config.yaml 中缺少 'api_key'。")
        sys.exit(1)
    except Exception as e:
        print(f"初始化富果API客戶端時發生錯誤：{e}")
        sys.exit(1)

def load_config(config_file):
    try:
        with open(config_file, 'r', encoding='utf-8') as f:
            return yaml.safe_load(f)
    except FileNotFoundError:
        print(f"錯誤：無法找到 {config_file} 文件。")
        sys.exit(1)
    except yaml.YAMLError as e:
        print(f"錯誤：讀取 {config_file} 文件時發生 YAML 錯誤：{e}")
        sys.exit(1)

def get_recent_trading_day():
    now = datetime.now()
    if now.time() >= datetime.strptime('13:31', '%H:%M').time():
        return now.strftime('%Y-%m-%d')
    else:
        return (now - timedelta(days=1)).strftime('%Y-%m-%d')

def calculate_5min_pct_increase_and_highest(candles_df):
    candles_df['MA5'] = candles_df['close'].rolling(window=5).mean()
    highest = 0.0
    for idx in range(len(candles_df)):
        current_time = candles_df.loc[idx, 'time']
        if highest == 0.0:
            highest = candles_df.loc[idx, 'high']
        else:
            highest = max(highest, candles_df.loc[idx, 'high'])
        candles_df.loc[idx, 'highest'] = highest
        if idx < 4:
            start_idx = 0
            end_idx = idx
        else:
            start_idx = idx - 4
            end_idx = idx
        close_window = candles_df.loc[start_idx:end_idx, 'close']
        max_close = close_window.max()
        min_close = close_window.min()
        if close_window.idxmax() < close_window.idxmin():
            pct_increase = -((max_close - min_close) / min_close) * 100
        else:
            pct_increase = ((max_close - min_close) / min_close) * 100
        if idx == 0:
            candles_df.loc[idx, '5min_pct_increase'] = 0
        else:
            candles_df.loc[idx, '5min_pct_increase'] = pct_increase
    for idx in range(5):
        if idx < len(candles_df) and pd.isna(candles_df.loc[idx, 'MA5']):
            candles_df.loc[idx, 'MA5'] = candles_df['close'][:idx + 1].mean()
    return candles_df


def load_all_kline_data():
    daily_kline_data = {}
    intraday_kline_data = {}

    if os.path.exists('daily_kline_data.json'):
        with open('daily_kline_data.json', 'r', encoding='utf-8') as f:
            daily_kline_data = json.load(f)

    if os.path.exists('intraday_kline_data.json'):
        with open('intraday_kline_data.json', 'r', encoding='utf-8') as f:
            intraday_kline_data = json.load(f)

    return daily_kline_data, intraday_kline_data

def fetch_intraday_data(client, symbol, trading_day, daily_kline_df):
    try:
        candles_response = client.stock.intraday.candles(
            symbol=symbol,
            oddLot=False,
            timeframe='1',
            _from=trading_day,
            to=trading_day
        )
        if candles_response and candles_response.get('data'):
            candles = candles_response['data']
            candles_df = pd.DataFrame(candles)
            candles_df['time'] = pd.to_datetime(candles_df['date']).dt.time
            candles_df['date'] = pd.to_datetime(trading_day).date()
            candles_df['symbol'] = symbol

            candles_df['昨日收盤價'] = 0.0
            candles_df['漲停價'] = 0.0
            candles_df['5min_pct_increase'] = 0.0

            now = datetime.now()
            if now.weekday() < 5 and datetime.strptime('13:31', '%H:%M').time() <= now.time() <= datetime.strptime('15:00', '%H:%M').time():
                yesterday_close_price = daily_kline_df['close'].iloc[0]
            else:
                if len(daily_kline_df) > 1:
                    yesterday_close_price = daily_kline_df['close'].iloc[1]
                else:
                    print(f"{symbol} 的日 K 線資料不足，無法取得昨日收盤價")
                    return pd.DataFrame()

            candles_df['昨日收盤價'] = yesterday_close_price
            candles_df['漲停價'] = calculate_limit_up_price(yesterday_close_price)

            full_time_index = pd.date_range(start='09:00', end='13:30', freq='1min').time
            full_index = pd.MultiIndex.from_product([candles_df['date'].unique(), full_time_index], names=['date', 'time'])
            candles_df.set_index(['date', 'time'], inplace=True)
            candles_df = candles_df.reindex(full_index)

            candles_df[['symbol', '昨日收盤價', '漲停價']] = candles_df[['symbol', '昨日收盤價', '漲停價']].ffill().bfill()
            candles_df['close'] = candles_df['close'].ffill()
            candles_df['close'] = candles_df['close'].fillna(candles_df['昨日收盤價'])
            candles_df['open'] = candles_df['open'].ffill()
            candles_df['open'] = candles_df['open'].fillna(candles_df['close'])
            candles_df['high'] = candles_df['high'].ffill()
            candles_df['high'] = candles_df['high'].fillna(candles_df['close'])
            candles_df['low'] = candles_df['low'].ffill()
            candles_df['low'] = candles_df['low'].fillna(candles_df['close'])
            candles_df['volume'] = candles_df['volume'].fillna(0)

            if '5min_pct_increase' not in candles_df.columns:
                candles_df['5min_pct_increase'] = 0.0
            else:
                candles_df['5min_pct_increase'] = candles_df['5min_pct_increase'].fillna(0.0)

            candles_df.reset_index(inplace=True)

            candles_df['rise'] = (candles_df['close'] - candles_df['昨日收盤價']) / candles_df['昨日收盤價'] * 100
            candles_df = calculate_5min_pct_increase_and_highest(candles_df)
            candles_df['highest'] = candles_df['highest'].ffill().bfill()
            if 'average' in candles_df.columns:
                candles_df = candles_df.drop(columns=['average'])

            print(f"{symbol} 的一分K數據已成功處理並返回。")
            return candles_df

    except Exception as e:
        print(f"獲取 {symbol} 在 {trading_day} 的一分K數據時出錯：{e}")
        return pd.DataFrame()

    return pd.DataFrame()

def get_recent_trading_day():
    today = datetime.now().date()
    weekday = today.weekday()
    if weekday >= 5:
        offset = weekday - 4
        recent_trading_day = today - timedelta(days=offset)
    else:
        recent_trading_day = today
    return recent_trading_day

def fetch_daily_kline_data(client, symbol, days=2):
    """
    獲取指定股票的日K線數據
    """
    end_date = get_recent_trading_day()
    start_date = end_date - timedelta(days=days)
    start_date_str = start_date.strftime('%Y-%m-%d')
    end_date_str = end_date.strftime('%Y-%m-%d')

    print(f"正在取得 {symbol} 從 {start_date_str} 到 {end_date_str} 的日K數據...")

    try:
        data = client.stock.historical.candles(symbol=symbol, from_=start_date_str, to=end_date_str)
        if 'data' in data and data['data']:
            daily_kline_df = pd.DataFrame(data['data'])
            return daily_kline_df
        else:
            print(f"無法取得 {symbol} 的日K數據：API 回應中不包含 'data' 欄位或 'data' 為空")
            return pd.DataFrame()
    except Exception as e:
        print(f"無法取得 {symbol} 的日K數據：{e}")
        return pd.DataFrame()

def save_matrix_dict(matrix_dict):
    with open('matrix_dict_analysis.json', 'w', encoding='utf-8') as f:
        json.dump(matrix_dict, f, indent=4, ensure_ascii=False)

def load_matrix_dict_analysis():
    if os.path.exists('matrix_dict_analysis.json'):
        with open('matrix_dict_analysis.json', 'r', encoding='utf-8') as f:
            return json.load(f)
    else:
        print("matrix_dict_analysis.json 文件不存在。")
        return {}

def filter_and_save_similarity_results(similarity_df):
    filtered_results = similarity_df[similarity_df['similarity_score'] >= 0.3]
    if not filtered_results.empty:
        mt_matrix_dict = filtered_results.to_dict(orient="records")
        save_mt_matrix_dict(mt_matrix_dict)
    else:
        print("沒有相似度大於等於 0.3 的結果，mt_matrix_dict.json 未建立")

def calculate_similarity_and_save(symbols_to_analyze):
    daily_kline_data, intraday_kline_data = load_kline_data()
    
    stock_data_list = []
    for symbol in symbols_to_analyze:
        if symbol not in daily_kline_data or symbol not in intraday_kline_data:
            print(f"無法取得 {symbol} 的日 K 線或一分 K 線數據，跳過。")
            continue
        
        daily_kline_df = pd.DataFrame(daily_kline_data[symbol])
        intraday_data = pd.DataFrame(intraday_kline_data[symbol])

        if not intraday_data.empty:
            stock_data_list.append(intraday_data)

    if stock_data_list:
        similarity_df = calculate_kline_similarity(stock_data_list)
        if not similarity_df.empty:
            save_mt_matrix_dict(similarity_df.to_dict(orient="records"))
            print("相似度分析結果已儲存")
        else:
            print("沒有相似的股票對，mt_matrix_dict.json 未建立")
    else:
        print("無法獲取有效的一分K資料，無法進行相似度分析")

def save_di_matrix_dict(di_matrix_dict):
    with open('di_matrix_dict.json', 'w', encoding='utf-8') as f:
        json.dump(di_matrix_dict, f, indent=4, ensure_ascii=False)

def load_di_matrix_dict():
    if os.path.exists('di_matrix_dict.json'):
        with open('di_matrix_dict.json', 'r', encoding='utf-8') as f:
            return json.load(f)
    else:
        return {}

def save_nb_matrix_dict(nb_matrix_dict):
    with open('nb_matrix_dict.json', 'w', encoding='utf-8') as f:
        json.dump(nb_matrix_dict, f, indent=4, ensure_ascii=False, default=str)

def merge_all_stock_data(stock_data_collection):
    merged_df = None
    for symbol, df in stock_data_collection.items():
        if not isinstance(df, pd.DataFrame):
            print(f"股票代號 {symbol} 的資料不是 DataFrame，跳過。")
            continue
        required_columns = ['time', '5min_pct_increase', 'rise']
        if not all(col in df.columns for col in required_columns):
            print(f"股票代號 {symbol} 的資料缺少必要欄位，跳過。")
            continue
        df_selected = df[['time', '5min_pct_increase', 'rise']].copy()
        df_selected = df_selected.rename(columns={
            '5min_pct_increase': f'5min_pct_increase_{symbol}',
            'rise': f'rise_{symbol}'
        })
        if merged_df is None:
            merged_df = df_selected
        else:
            merged_df = pd.merge(merged_df, df_selected, on='time', how='outer')
    if merged_df is not None:
        merged_df = merged_df.sort_values('time').reset_index(drop=True)
    else:
        merged_df = pd.DataFrame()
    return merged_df

def initialize_stock_data(symbols_to_analyze, daily_kline_data, intraday_kline_data):
    stock_data_collection = {}
    for symbol in symbols_to_analyze:
        if symbol not in daily_kline_data or symbol not in intraday_kline_data:
            print(f"股票代號 {symbol} 的日 K 線或一分 K 線資料缺失，跳過。")
            continue
        daily_kline_df = pd.DataFrame(daily_kline_data[symbol])
        intraday_data = pd.DataFrame(intraday_kline_data[symbol])
        if intraday_data.empty:
            print(f"股票代號 {symbol} 的日內數據為空，跳過。")
            continue
        complete_df = ensure_continuous_time_series(intraday_data)
        complete_df = complete_df.drop(columns=['volume', 'average'], errors='ignore')
        stock_data_collection[symbol] = complete_df
    return stock_data_collection

def process_group_data(stock_data_collection, wait_minutes, hold_minutes, matrix_dict_analysis, verbose=True):
    global capital_per_stock, transaction_fee, transaction_discount, trading_tax
    global price_gap_below_50, price_gap_50_to_100, price_gap_100_to_500, price_gap_500_to_1000, price_gap_above_1000
    global allow_reentry_after_stop_loss

    merged_df = None
    for symbol, df in stock_data_collection.items():
        if not isinstance(df, pd.DataFrame):
            print(f"股票代號 {symbol} 的數據不是 DataFrame，跳過。")
            continue
        required_columns = ['time', '5min_pct_increase', 'rise', 'high']
        if not all(col in df.columns for col in required_columns):
            print(f"股票代號 {symbol} 的資料缺少必要列，跳過。")
            continue
        df_selected = df[['time', '5min_pct_increase', 'rise', 'high']].copy()
        df_selected = df_selected.rename(columns={
            '5min_pct_increase': f'5min_pct_increase_{symbol}',
            'rise': f'rise_{symbol}',
            'high': f'high_{symbol}'
        })
        if merged_df is None:
            merged_df = df_selected
        else:
            merged_df = pd.merge(merged_df, df_selected, on='time', how='outer')
    if merged_df is not None:
        merged_df = merged_df.sort_values('time').reset_index(drop=True)
    else:
        merged_df = pd.DataFrame()

    total_bars = len(merged_df)
    merged_df_list = list(merged_df.iterrows())
    idx = 0
    in_position = False
    has_exited = False
    stop_loss_triggered = False
    total_profit = 0
    total_profit_rate = 0
    total_trades = 0
    message_log = []
    stock_symbols = list(stock_data_collection.keys())
    previous_high_values = {}
    previous_rise_values = {}
    already_entered_stocks = []
    final_check_active = False
    final_check_count = 0
    final_check_max = 10
    can_trade = True
    price_gap_below_50 = below_50
    already_triggered_limit_up = set()
    leader = None
    tracking_stocks = set()
    leader_rise_before_decline = None
    in_waiting_period = False
    waiting_time = 0
    hold_time = 0
    first_condition_one_time = None
    leader_peak_rise = None
    backtrack = False

    def check_5min_pct_increase(stock, start_time, end_time):
        stock_df = stock_data_collection.get(stock, pd.DataFrame())
        if stock_df.empty:
            return False
        period_data = stock_df[(stock_df['time'] >= start_time) & (stock_df['time'] <= end_time)]
        return period_data['5min_pct_increase'].gt(1).any()

    def check_high_values_during_period(stock, start_time, end_time):
        stock_df = stock_data_collection.get(stock, pd.DataFrame())
        if stock_df.empty:
            return False
        period_data = stock_df[(stock_df['time'] >= start_time) & (stock_df['time'] <= end_time)]
        period_data = period_data.sort_values(by='time').reset_index(drop=True)
        for i in range(1, len(period_data)):
            if period_data.loc[i, 'high'] <= period_data.loc[i - 1, 'high']:
                return True
        return False

    while idx < total_bars:
        index, row = merged_df_list[idx]
        current_time = row['time']
        current_time_str = current_time.strftime('%H:%M:%S')

        for symbol in stock_symbols:
            stock_df = stock_data_collection[symbol]
            current_row = stock_df[stock_df['time'] == current_time]
            if not current_row.empty:
                rise_col = f'rise_{symbol}'
                row[rise_col] = current_row['rise'].values[0]
                high_col = f'high_{symbol}'
                row[high_col] = current_row['high'].values[0]
                five_min_pct_increase_col = f'5min_pct_increase_{symbol}'
                row[five_min_pct_increase_col] = current_row['5min_pct_increase'].values[0]
            else:
                row[f'rise_{symbol}'] = None
                row[f'high_{symbol}'] = None

        if in_position and not has_exited:
            hold_time += 1
            if hold_minutes is not None:
                if hold_time >= hold_minutes:
                    profit, profit_rate = exit_trade(
                        stock_data_collection[current_position['symbol']],
                        current_position['shares'],
                        current_position['entry_price'],
                        current_position['sell_cost'],
                        current_position['entry_fee'],
                        current_position['tax'],
                        message_log,
                        current_time,
                        hold_time,
                        current_position['entry_time']
                    )
                    total_trades += 1
                    total_profit += profit
                    total_profit_rate += profit_rate
                    in_position = False
                    has_exited = True
                    current_position = None
            elif current_time.strftime('%H:%M:%S') == '13:30:00':
                profit, profit_rate = exit_trade(
                    stock_data_collection[current_position['symbol']],
                    current_position['shares'],
                    current_position['entry_price'],
                    current_position['sell_cost'],
                    current_position['entry_fee'],
                    current_position['tax'],
                    message_log,
                    current_time,
                    hold_time,
                    current_position['entry_time'],
                    use_f_exit=True
                )
                total_trades += 1
                total_profit += profit
                total_profit_rate += profit_rate
                in_position = False
                has_exited = True
                current_position = None
                idx += 1
                continue

        if in_position and not has_exited:
            selected_symbol = current_position['symbol']
            selected_stock_df = stock_data_collection[selected_symbol]
            current_row = selected_stock_df[selected_stock_df['time'] == current_time]
            if not current_row.empty:
                current_high = current_row['high'].values[0]
                price_difference = (current_position['highest_on_entry'] - current_position['entry_price']) * 1000

                if price_difference < current_position['current_price_gap']:
                    stop_loss_type = 'price_difference'
                    stop_loss_threshold = current_position['entry_price'] + (current_position['current_price_gap'] / 1000)
                else:
                    stop_loss_type = 'over_high'
                    stop_loss_threshold = current_position['highest_on_entry'] + current_position['tick_unit']

                if current_high >= stop_loss_threshold:
                    exit_price = stop_loss_threshold
                    exit_reason = f"條件三觸發{stop_loss_type}停損）"
                    trigger_exit = True
                else:
                    trigger_exit = False

                if trigger_exit:
                    exit_cost = current_position['shares'] * exit_price * 1000
                    exit_fee = int(exit_cost * (transaction_fee * 0.01) * (transaction_discount * 0.01))
                    profit = current_position['sell_cost'] - exit_cost - current_position['entry_fee'] - exit_fee - current_position['tax']
                    profit_rate = profit * 100 / current_position['sell_cost'] if current_position['sell_cost'] != 0 else 0.0
                    message_log.append(
                        (current_time_str,
                         f"{RED}{exit_reason}！出場成功！{RESET}")
                    )
                    message_log.append(
                        (current_time_str,
                         f"{RED}股票代號：{current_position['symbol']}，持有張數：{current_position['shares']} 張，出場價格：{exit_price} 元，出場價金：{int(exit_cost)} 元，利潤：{int(profit)} 元，報酬率：{profit_rate:.2f}%，手續費：{exit_fee} 元{RESET}")
                    )
                    total_trades += 1
                    total_profit += profit
                    total_profit_rate += profit_rate
                    in_position = False
                    has_exited = True
                    current_position = None
                    stop_loss_triggered = True

                    if allow_reentry_after_stop_loss:
                        backtrack_start_idx = max(0, idx - 5)
                        idx = backtrack_start_idx
                        backtrack = True
                        leader = None
                        tracking_stocks = set()
                        previous_rise_values = {}
                        leader_rise_before_decline = None
                        in_waiting_period = False
                        waiting_time = 0
                        already_entered_stocks = []
                        final_check_active = False
                        final_check_count = 0
                        can_trade = True
                        hold_time = 0
                        first_condition_one_time = None
                        leader_peak_rise = None
                        if verbose:
                            message_log.append(
                                (current_time_str, f"{YELLOW}[回朔] 觸發條件三，回溯五根K棒至 {merged_df_list[idx][1]['time'].strftime('%H:%M:%S')}，檢查是否有新的進場機會{RESET}")
                            )
                        continue
                    else:
                        message_log.append((current_time_str, "停損後無其它進場機會，結束程序"))
                        break
            else:
                pass

        if in_position:
            idx += 1
            continue

        for symbol in stock_symbols:
            stock_df = stock_data_collection[symbol]
            current_row = stock_df[stock_df['time'] == current_time]
            if current_row.empty:
                continue

            current_high = current_row['high'].values[0]
            limit_up_price = current_row['漲停價'].values[0]

            pct_increase = current_row['5min_pct_increase'].values[0]
            rise = current_row['rise'].values[0]

            if current_high == limit_up_price and symbol not in tracking_stocks:
                tracking_stocks.add(symbol)
                if verbose and not in_waiting_period and not final_check_active:
                    message_log.append(
                        (current_time_str, f"領漲 {symbol} 漲停")
                    )
                leader = symbol
                in_waiting_period = True
                waiting_time = 1
                break

            if current_high != limit_up_price and pct_increase >= 1.5 and symbol not in tracking_stocks:
                tracking_stocks.add(symbol)
                if verbose and not in_waiting_period and not final_check_active:
                    message_log.append(
                        (current_time_str, f"股票代號:{symbol} 觸發條件一，rise: {rise:.2f}%")
                    )
                first_condition_one_time = current_time

        if tracking_stocks:
            max_rise = None
            new_leader = leader
            for symbol in tracking_stocks:
                rise = row.get(f'rise_{symbol}', None)
                if rise is not None:
                    if max_rise is None or rise > max_rise:
                        max_rise = rise
                        new_leader = symbol
            if new_leader != leader:
                if verbose and leader is not None:
                    message_log.append(
                        (current_time_str, f"領漲者變更為 {new_leader}，rise: {max_rise:.2f}%")
                    )
                leader = new_leader
                leader_peak_rise = max_rise

                if in_waiting_period:
                    in_waiting_period = False
                    waiting_time = 0
                    if verbose:
                        message_log.append(
                            (current_time_str, f"領漲變更，重置等待時間")
                        )

            if leader and not in_waiting_period and not final_check_active:
                rise = row.get(f'rise_{leader}', None)
                if verbose and rise is not None:
                    message_log.append(
                        (current_time_str, f"領漲 {leader}，rise: {rise:.2f}%")
                    )

            current_rise = row.get(f'rise_{leader}', None)
            prev_rise = previous_rise_values.get(leader)

            if not final_check_active:
                if prev_rise is not None and current_rise is not None:
                    if current_rise <= prev_rise:
                        if not in_waiting_period:
                            in_waiting_period = True
                            waiting_time = 1
                            previous_time = (datetime.combine(date.today(), current_time) - timedelta(minutes=1)).time()
                            previous_rise_value_series = stock_data_collection[leader][stock_data_collection[leader]['time'] == previous_time]['rise']
                            if not previous_rise_value_series.empty:
                                leader_rise_before_decline = previous_rise_value_series.values[0]
                            else:
                                leader_rise_before_decline = current_rise
                previous_rise_values[leader] = current_rise

        else:
            idx += 1
            continue

        if in_waiting_period:
            for symbol in tracking_stocks:
                if symbol == leader:
                    continue
                rise = row.get(f'rise_{symbol}', None)
                if rise is not None and leader_rise_before_decline is not None:
                    if rise > leader_rise_before_decline:
                        final_check_active = False
                        final_check_count = 0
                        in_waiting_period = False
                        waiting_time = 0
                        leader_peak_rise = rise
                        if verbose:
                            message_log.append(
                                (current_time_str, f"領漲 {leader} 超越記錄的 rise 值，重置流程")
                            )
                else:
                    leader_rise_before_decline = rise

            if in_waiting_period:
                if verbose:
                    message_log.append(
                        (current_time_str,
                         f"等待中，第 {waiting_time} 分鐘")
                    )
                if waiting_time >= wait_minutes:
                    in_waiting_period = False
                    waiting_time = 0
                    final_check_active = False
                    final_check_count = 0
                    if verbose:
                        message_log.append(
                            (current_time_str,
                             "等待完成，開始檢查是否有符合進場條件的股票")
                        )
                    eligible_stocks = []
                    group_name = None
                    for group, symbols in matrix_dict_analysis.items():
                        if leader in symbols:
                            group_name = group
                            break

                    if group_name is None:
                        print(f"無法找到領漲 {leader} 所屬的族群，無法進行檢查。")
                        idx += 1
                        continue

                    nb_matrix_dict = load_nb_matrix_dict()
                    consolidated_symbols = nb_matrix_dict.get('consolidated_symbols', {})
                    if group_name in consolidated_symbols:
                        nb_symbols = consolidated_symbols[group_name]
                        for symbol in nb_symbols:
                            if symbol == leader:
                                continue
                            stock_df = stock_data_collection.get(symbol, pd.DataFrame())
                            if stock_df.empty:
                                continue

                            has_pct_increase = check_5min_pct_increase(symbol, first_condition_one_time, current_time)
                            if not has_pct_increase:
                                continue 

                            has_high_decrease = check_high_values_during_period(symbol, first_condition_one_time, current_time)
                            if not has_high_decrease:
                                continue

                            current_rise = row.get(f'rise_{symbol}', 0)
                            if not (-3 < current_rise < 8):
                                continue

                            eligible_stocks.append({
                                'symbol': symbol,
                                'rise': current_rise
                            })
                    else:
                        print(f"{group_name} 不在 consolidated_symbols 中")

                    if eligible_stocks:
                        eligible_stocks_sorted = sorted(eligible_stocks, key=lambda x: x['rise'], reverse=True)
                        median_index = len(eligible_stocks_sorted) // 2
                        selected_stock = eligible_stocks_sorted[median_index]
                        selected_symbol = selected_stock['symbol']
                        selected_stock_rise = selected_stock['rise']
                        entry_price_series = stock_data_collection[selected_symbol][stock_data_collection[selected_symbol]['time'] == current_time]['close']

                        if not entry_price_series.empty:
                            entry_price = entry_price_series.values[0]
                            shares = round((capital_per_stock * 10000) / (entry_price * 1000))
                            sell_cost = shares * entry_price * 1000
                            entry_fee = int(sell_cost * (transaction_fee * 0.01) * (transaction_discount * 0.01))
                            tax = int(sell_cost * (trading_tax * 0.01))
                            if entry_price < 10:
                                current_price_gap = price_gap_below_50
                                tick_unit = 0.01
                            elif entry_price < 50:
                                current_price_gap = price_gap_below_50
                                tick_unit = 0.05
                            elif entry_price < 100:
                                current_price_gap = price_gap_50_to_100
                                tick_unit = 0.1
                            elif entry_price < 500:
                                current_price_gap = price_gap_100_to_500
                                tick_unit = 0.5
                            elif entry_price < 1000:
                                current_price_gap = price_gap_500_to_1000
                                tick_unit = 1
                            else:
                                current_price_gap = price_gap_above_1000
                                tick_unit = 5

                            current_position = {
                                'symbol': selected_symbol,
                                'shares': shares,
                                'entry_price': entry_price,
                                'sell_cost': sell_cost,
                                'entry_fee': entry_fee,
                                'tax': tax,
                                'entry_time': current_time_str,
                                'entry_index': idx,
                                'current_price_gap': current_price_gap,
                                'tick_unit': tick_unit,
                                'highest_on_entry': stock_data_collection[selected_symbol][stock_data_collection[selected_symbol]['time'] == current_time]['highest'].values[0],
                                'initial_highest': stock_data_collection[selected_symbol][stock_data_collection[selected_symbol]['time'] == current_time]['highest'].values[0],
                                'stop_loss_type': None,
                                'stop_loss_threshold': None
                            }

                            message_log.append(
                                (current_time_str,
                                f"{GREEN}進場！股票代號：{selected_symbol}，進場 {shares} 張，進場價格：{entry_price} 元，進場價金：{int(sell_cost)} 元，手續費：{entry_fee} 元，證交稅：{tax} 元。{RESET}")
                            )

                            in_position = True
                            has_exited = False
                            already_entered_stocks.append(selected_symbol)
                            hold_time = 0
                            can_trade = False
                            if allow_reentry_after_stop_loss:
                                stop_loss_triggered = False

                            price_difference = (current_position['highest_on_entry'] - current_position['entry_price']) * 1000
                            if price_difference < current_position['current_price_gap']:
                                current_position['stop_loss_type'] = 'price_difference'
                                current_position['stop_loss_threshold'] = current_position['entry_price'] + (current_position['current_price_gap'] / 1000)
                            else:
                                current_position['stop_loss_type'] = 'over_high'
                                current_position['stop_loss_threshold'] = current_position['highest_on_entry'] + current_position['tick_unit']

                            final_check_active = False
                            final_check_count = 0
                            in_waiting_period = False
                            waiting_time = 0
                            hold_time = 0
                            leader = None
                            tracking_stocks = set()
                            previous_rise_values = {}
                            leader_peak_rise = None
                            leader_rise_before_decline = None
                            first_condition_one_time = None
                            idx += 1
                            continue
                        else:
                            message_log.append(
                                (current_time_str,
                                 f"無法取得 {selected_symbol} 在 {current_time_str} 的價格，進場失敗")
                            )
                            idx += 1
                            continue
                    else:
                        final_check_active = True
                        final_check_count = 0
                        if verbose:
                            message_log.append(
                                (current_time_str,
                                 "沒有符合進場條件的股票，進入最後十次檢查階段")
                            )
                else:
                    waiting_time += 1
                idx += 1
                continue

        if final_check_active:
            final_check_count += 1
            if verbose:
                message_log.append(
                    (current_time_str,
                    f"最後檢查第 {final_check_count} 分鐘")
                )

            if leader and row.get(f'high_{leader}', None) == row.get('漲停價', None):
                continue

            rise = row.get(f'rise_{leader}', None)
            if rise is not None and leader_rise_before_decline is not None and rise > leader_rise_before_decline:
                final_check_active = False
                final_check_count = 0
                in_waiting_period = False
                waiting_time = 0
                leader_peak_rise = rise
                if verbose:
                    message_log.append(
                        (current_time_str, f"領漲 {leader} 超越記錄的 rise 值，重置流程")
                    )
                idx += 1
                continue

            eligible_stocks = []
            group_name = None
            for group, symbols in matrix_dict_analysis.items():
                if leader in symbols:
                    group_name = group
                    break

            if group_name is None:
                print(f"無法找到領漲 {leader} 所屬的族群，無法進行檢查。")
                idx += 1
                continue

            nb_matrix_dict = load_nb_matrix_dict()
            if group_name in nb_matrix_dict:
                nb_symbols = nb_matrix_dict[group_name]
                for symbol in nb_symbols:
                    if symbol == leader:
                        continue

                    stock_df = stock_data_collection.get(symbol, pd.DataFrame())
                    if stock_df.empty:
                        continue

                    has_pct_increase = check_5min_pct_increase(symbol, first_condition_one_time, current_time)
                    if not has_pct_increase:
                        continue

                    has_high_decrease = check_high_values_during_period(symbol, first_condition_one_time, current_time)
                    if not has_high_decrease:
                        continue

                    current_rise = row.get(f'rise_{symbol}', 0)
                    if not (-3 < current_rise < 8):
                        continue

                    eligible_stocks.append({
                        'symbol': symbol,
                        'rise': current_rise
                    })

            if eligible_stocks:
                eligible_stocks_sorted = sorted(eligible_stocks, key=lambda x: x['rise'], reverse=True)
                median_index = len(eligible_stocks_sorted) // 2
                selected_stock = eligible_stocks_sorted[median_index]
                selected_symbol = selected_stock['symbol']
                selected_stock_rise = selected_stock['rise']
                entry_price_series = stock_data_collection[selected_symbol][stock_data_collection[selected_symbol]['time'] == current_time]['close']

                if not entry_price_series.empty:
                    entry_price = entry_price_series.values[0]
                    shares = round((capital_per_stock * 10000) / (entry_price * 1000))
                    sell_cost = shares * entry_price * 1000
                    entry_fee = int(sell_cost * (transaction_fee * 0.01) * (transaction_discount * 0.01))
                    tax = int(sell_cost * (trading_tax * 0.01))
                    if entry_price < 10:
                        current_price_gap = price_gap_below_50
                        tick_unit = 0.01
                    elif entry_price < 50:
                        current_price_gap = price_gap_below_50
                        tick_unit = 0.05
                    elif entry_price < 100:
                        current_price_gap = price_gap_50_to_100
                        tick_unit = 0.1
                    elif entry_price < 500:
                        current_price_gap = price_gap_100_to_500
                        tick_unit = 0.5
                    elif entry_price < 1000:
                        current_price_gap = price_gap_500_to_1000
                        tick_unit = 1
                    else:
                        current_price_gap = price_gap_above_1000
                        tick_unit = 5

                    current_position = {
                        'symbol': selected_symbol,
                        'shares': shares,
                        'entry_price': entry_price,
                        'sell_cost': sell_cost,
                        'entry_fee': entry_fee,
                        'tax': tax,
                        'entry_time': current_time_str,
                        'entry_index': idx,
                        'current_price_gap': current_price_gap,
                        'tick_unit': tick_unit,
                        'highest_on_entry': stock_data_collection[selected_symbol][stock_data_collection[selected_symbol]['time'] == current_time]['highest'].values[0],
                        'initial_highest': stock_data_collection[selected_symbol][stock_data_collection[selected_symbol]['time'] == current_time]['highest'].values[0],
                        'stop_loss_type': None,
                        'stop_loss_threshold': None
                    }

                    message_log.append(
                        (current_time_str,
                        f"{GREEN}進場！股票代號：{selected_symbol}，進場 {shares} 張，進場價格：{entry_price} 元，進場價金：{int(sell_cost)} 元，手續費：{entry_fee} 元，證交稅：{tax} 元。{RESET}")
                    )

                    in_position = True
                    has_exited = False
                    already_entered_stocks.append(selected_symbol)
                    hold_time = 0
                    can_trade = False
                    if allow_reentry_after_stop_loss:
                        stop_loss_triggered = False

                    price_difference = (current_position['highest_on_entry'] - current_position['entry_price']) * 1000
                    if price_difference < current_position['current_price_gap']:
                        current_position['stop_loss_type'] = 'price_difference'
                        current_position['stop_loss_threshold'] = current_position['entry_price'] + (current_position['current_price_gap'] / 1000)
                    else:
                        current_position['stop_loss_type'] = 'over_high'
                        current_position['stop_loss_threshold'] = current_position['highest_on_entry'] + current_position['tick_unit']

                    final_check_active = False
                    final_check_count = 0
                    in_waiting_period = False
                    waiting_time = 0
                    hold_time = 0
                    leader = None
                    tracking_stocks = set()
                    previous_rise_values = {}
                    leader_peak_rise = None
                    leader_rise_before_decline = None
                    first_condition_one_time = None
                    idx += 1
                    continue
                else:
                    message_log.append(
                        (current_time_str,
                         f"無法取得 {selected_symbol} 在 {current_time_str} 的價格，進場失敗")
                    )
                    idx += 1
                    continue
            else:
                if final_check_count >= final_check_max:
                    if verbose:
                        message_log.append(
                            (current_time_str,
                             f"{YELLOW}最後檢查完成，仍未發現可進場股票{RESET}")
                        )

                    final_check_active = False
                    final_check_count = 0
                    in_waiting_period = False
                    waiting_time = 0
                    hold_time = 0
                    leader = None
                    tracking_stocks = set()
                    previous_rise_values = {}
                    leader_peak_rise = None
                    leader_rise_before_decline = None
                    first_condition_one_time = None
                idx += 1
                continue

        idx += 1

    message_log.sort(key=lambda x: str(x[0]))
    for log_time, message in message_log:
        print(f"[{log_time}] {message}")

    if total_trades > 0:
        avg_profit_rate = total_profit_rate / total_trades
        return total_profit, avg_profit_rate
    else:
        if verbose:
            print("無交易，無法計算總利潤和報酬率")
        return None, None

def exit_trade(
    selected_stock_df, shares, entry_price, sell_cost,
    entry_fee, tax,
    message_log, current_time, hold_time, entry_time, use_f_exit=False
):
    current_time_str = current_time if isinstance(current_time, str) else current_time.strftime('%H:%M:%S')
    selected_stock_df['time'] = pd.to_datetime(selected_stock_df['time'], format='%H:%M:%S').dt.time
    if isinstance(entry_time, str):
        entry_time_obj = pd.to_datetime(entry_time, format='%H:%M:%S').time()
    else:
        entry_time_obj = entry_time
    if use_f_exit:
        end_price_series = selected_stock_df[selected_stock_df['time'] == datetime.strptime('13:30', '%H:%M').time()]['close']
        if not end_price_series.empty:
            end_price = end_price_series.values[0]
        else:
            print("無法取得 13:30 的數據，出場時間配對錯誤")
            message_log.append((current_time_str, "出場時間配對錯誤"))
            return None, None
    else:
        entry_row = selected_stock_df[selected_stock_df['time'] == entry_time_obj]
        if not entry_row.empty:
            entry_index = entry_row.index[0]
            exit_index = entry_index + hold_time
            if exit_index >= len(selected_stock_df):
                print("出場時間超出範圍，無法進行交易")
                message_log.append((current_time_str, "出場時間超出範圍"))
                return None, None
            end_price = selected_stock_df.iloc[exit_index]['close']
        else:
            print("進場時間配對錯誤，無法找到精確的進場時間")
            message_log.append((current_time_str, "進場時間配對錯誤"))
            return None, None
    buy_cost = shares * end_price * 1000
    exit_fee = int(buy_cost * transaction_fee * 0.01 * transaction_discount * 0.01)
    profit = sell_cost - buy_cost - entry_fee - exit_fee - tax
    return_rate = (profit * 100) / (buy_cost - exit_fee) if (buy_cost - exit_fee) != 0 else 0.0
    if use_f_exit:
        message_log.append(
            (current_time_str,
             f"{RED}股票出場，持有時間 {hold_time} 分鐘（強制出場）{RESET}")
        )
    else:
        message_log.append(
            (current_time_str,
             f"{RED}股票出場，持有時間 {hold_time} 分鐘{RESET}")
        )
    message_log.append(
        (current_time_str,
         f"{RED}持有張數：{shares} 張，出場價格：{end_price} 元，出場價金：{int(buy_cost)} 元，利潤：{int(profit)} 元，報酬率：{return_rate:.2f}%，手續費：{exit_fee} 元{RESET}")
    )
    return profit, return_rate

def consolidate_and_save_stock_symbols():
    mt_matrix_dict = load_mt_matrix_dict()
    matrix_dict_analysis = load_matrix_dict_analysis()
    
    if not mt_matrix_dict:
        print("mt_matrix_dict.json 文件不存在或為空，無法進行統整")
        return
    if not matrix_dict_analysis:
        print("matrix_dict_analysis.json 文件不存在或為空，無法進行統整")
        return

    consolidated_group_symbols = {group: [] for group in matrix_dict_analysis.keys()}
    
    for record in mt_matrix_dict:
        stock1 = record.get('stock1')
        stock2 = record.get('stock2')
        similarity_score = record.get('similarity_score', 0)
        
        if similarity_score >= 0.3:
            for group, symbols in matrix_dict_analysis.items():
                if stock1 in symbols and stock1 not in consolidated_group_symbols[group]:
                    consolidated_group_symbols[group].append(stock1)
                if stock2 in symbols and stock2 not in consolidated_group_symbols[group]:
                    consolidated_group_symbols[group].append(stock2)
    
    for group in consolidated_group_symbols:
        consolidated_group_symbols[group] = list(set(consolidated_group_symbols[group]))
    
    nb_matrix_dict = {"consolidated_symbols": consolidated_group_symbols}
    
    save_nb_matrix_dict(nb_matrix_dict)
    print(f"統整後的股票代號已保存至 nb_matrix_dict.json，按族群分類。")

def load_and_filter_symbols():
    if os.path.exists('matrix_dict_analysis.json'):
        with open('matrix_dict_analysis.json', 'r', encoding='utf-8') as f:
            matrix_dict_analysis = json.load(f)
    else:
        print("matrix_dict_analysis.json 文件不存在。")
        return []
    disposition_stocks = load_disposition_stocks()
    all_symbols = []
    for group in matrix_dict_analysis.values():
        all_symbols.extend(group)
    filtered_symbols = []
    removed_stocks = []
    for symbol in all_symbols:
        if symbol in disposition_stocks:
            removed_stocks.append(symbol)
        else:
            filtered_symbols.append(symbol)
    if removed_stocks:
        for stock in removed_stocks:
            print(f"處置股 {stock} 已被剃除")
    else:
        print("沒有股票是處置股")
    return filtered_symbols

def calculate_kline_similarity(stock_data_list):
    similarity_results = []
    for i in range(len(stock_data_list)):
        stock1 = stock_data_list[i]
        if 'symbol' not in stock1.columns:
            raise KeyError("DataFrame does not contain 'symbol' column.")
        for j in range(i + 1, len(stock_data_list)):
            stock2 = stock_data_list[j]
            if 'symbol' not in stock2.columns:
                raise KeyError("DataFrame does not contain 'symbol' column.")
            symbol1 = stock1['symbol'].iloc[0]
            symbol2 = stock2['symbol'].iloc[0]
            if symbol1 != symbol2:
                merged_df = pd.merge(stock1, stock2, on='time', suffixes=('_1', '_2'))
                merged_df['昨日收盤價_2'] = merged_df['昨日收盤價_2'].ffill().bfill()
                if 'high_1' not in merged_df.columns or 'high_2' not in merged_df.columns:
                    print(f"股票 {symbol1} 或 {symbol2} 缺少 'high' 欄位，跳過相似度計算。")
                    continue
                for col in ['open', 'high', 'low', 'close']:
                    merged_df[f'{col}_1_z'] = (merged_df[f'{col}_1'] - merged_df[f'{col}_1'].mean()) / merged_df[f'{col}_1'].std()
                    merged_df[f'{col}_2_z'] = (merged_df[f'{col}_2'] - merged_df[f'{col}_2'].mean()) / merged_df[f'{col}_2'].std()
                distance = np.sqrt(
                    (merged_df['open_1_z'] - merged_df['open_2_z']) ** 2 +
                    (merged_df['high_1_z'] - merged_df['high_2_z']) ** 2 +
                    (merged_df['low_1_z'] - merged_df['low_2_z']) ** 2 +
                    (merged_df['close_1_z'] - merged_df['close_2_z']) ** 2
                ).mean()
                similarity_score = 1 / (1 + distance)
                if similarity_score >= 0.3:
                    result = {
                        'stock1': symbol1,
                        'stock2': symbol2,
                        'similarity_score': similarity_score
                    }
                    similarity_results.append(result)
    if not similarity_results:
        print("沒有找到相似度大於等於 0.3 的結果")
        return pd.DataFrame(columns=['stock1', 'stock2', 'similarity_score'])
    similarity_df = pd.DataFrame(similarity_results)
    similarity_df = similarity_df.sort_values(by='similarity_score', ascending=False).reset_index(drop=True)
    return similarity_df

def calculate_limit_up_price(close_price):
    limit_up = close_price * 1.10
    if limit_up < 10:
        price_unit = 0.01
    elif limit_up < 50:
        price_unit = 0.05
    elif limit_up < 100:
        price_unit = 0.1
    elif limit_up < 500:
        price_unit = 0.5
    elif limit_up < 1000:
        price_unit = 1
    else:
        price_unit = 5
    limit_up_price = (limit_up // price_unit) * price_unit
    return limit_up_price

def save_mt_matrix_dict(mt_matrix_dict):
    with open('mt_matrix_dict.json', 'w', encoding='utf-8') as f:
        json.dump(mt_matrix_dict, f, indent=4, ensure_ascii=False, default=str)

def load_mt_matrix_dict():
    if os.path.exists('mt_matrix_dict.json'):
        with open('mt_matrix_dict.json', 'r', encoding='utf-8') as f:
            return json.load(f)
    else:
        return {}

def load_nb_matrix_dict():
    if os.path.exists('nb_matrix_dict.json'):
        with open('nb_matrix_dict.json', 'r', encoding='utf-8') as f:
            return json.load(f)
    else:
        return {}
    
def ensure_continuous_time_series(df):
    df['date'] = pd.to_datetime(df['date'])
    df['time'] = pd.to_datetime(df['time'], format='%H:%M:%S').dt.time

    full_time_index = pd.date_range(start='09:00', end='13:30', freq='1min').time

    full_index = pd.MultiIndex.from_product([df['date'].unique(), full_time_index], names=['date', 'time'])

    df.set_index(['date', 'time'], inplace=True)

    df = df.reindex(full_index)

    df[['symbol', '昨日收盤價', '漲停價']] = df[['symbol', '昨日收盤價', '漲停價']].ffill().bfill()

    if 'high' not in df.columns:
        df['high'] = df['close']
    if 'low' not in df.columns:
        df['low'] = df['close']

    df['close'] = df['close'].ffill()
    df['close'] = df['close'].fillna(df['昨日收盤價'])
    df['open'] = df['open'].ffill()
    df['open'] = df['open'].fillna(df['close'])
    df['high'] = df['high'].ffill()
    df['high'] = df['high'].fillna(df['close'])
    df['low'] = df['low'].ffill()
    df['low'] = df['low'].fillna(df['close'])

    df['volume'] = df['volume'].fillna(0)

    if '5min_pct_increase' not in df.columns:
        df['5min_pct_increase'] = 0.0
    else:
        df['5min_pct_increase'] = df['5min_pct_increase'].fillna(0.0)

    df.reset_index(inplace=True)

    df = calculate_5min_pct_increase_and_highest(df)

    return df

def print_and_complete_nb_matrix_dict():
    daily_kline_data, intraday_kline_data = load_kline_data()
    recent_day = get_recent_trading_day()
    nb_matrix_dict = load_nb_matrix_dict()
    consolidated_symbols = nb_matrix_dict.get("consolidated_symbols", [])
    disposition_stocks = load_disposition_stocks()
    
    filtered_symbols = [symbol for symbol in consolidated_symbols if symbol not in disposition_stocks]

    for symbol in filtered_symbols:
        print(f"\n股票代號：{symbol} 的完整數據：")
        if symbol not in daily_kline_data or symbol not in intraday_kline_data:
            print(f"無法取得 {symbol} 的日 K 線或一分 K 線數據，跳過。")
            continue

        daily_kline_df = pd.DataFrame(daily_kline_data[symbol])
        intraday_data = pd.DataFrame(intraday_kline_data[symbol])

        if not intraday_data.empty:
            complete_df = ensure_continuous_time_series(intraday_data)
            complete_df = complete_df.drop(columns=['volume', 'average'], errors='ignore')
            pd.set_option('display.max_rows', None)
            pd.set_option('display.max_columns', None)
            pd.set_option('display.width', 1000)
            print(complete_df)
            pd.reset_option('display.max_rows')
            pd.reset_option('display.max_columns')
            pd.reset_option('display.width')
        else:
            print(f"無法取得 {symbol} 的數據")

def save_disposition_stocks(disposition_stocks):
    """儲存處置股清單"""
    with open('Disposition.json', 'w', encoding='utf-8') as f:
        json.dump(disposition_stocks, f, indent=4, ensure_ascii=False)
        
def load_disposition_stocks():
    disposition_file = 'Disposition.json'
    try:
        with open(disposition_file, 'r', encoding='utf-8') as f:
            disposition_data = json.load(f)
            return disposition_data
    except FileNotFoundError:
        print(f"錯誤：無法找到 {disposition_file} 文件。")
        return []
    except json.JSONDecodeError:
        print(f"錯誤：{disposition_file} 文件格式不正確。")
        return []
    
def fetch_disposition_stocks(client, matrix_dict_analysis):
    disposition_stocks = []
    for group, stock_list in matrix_dict_analysis.items():
        for symbol in stock_list:
            try:
                ticker_data = client.stock.intraday.ticker(symbol=symbol)
                if ticker_data.get('isDisposition', False):
                    disposition_stocks.append(symbol)
            except Exception as e:
                print(f"獲取 {symbol} 的處置股狀態時發生錯誤: {e}")
    with open('Disposition.json', 'w', encoding='utf-8') as f:
        json.dump(disposition_stocks, f, indent=4, ensure_ascii=False)

def calculate_average_over_high_list():
    while True:
        print("\n選擇計算平均過高的模式：")
        print("1. 單一族群分析")
        print("2. 全部族群分析")
        print("0. 返回主選單")
        sub_choice = input("請輸入選項：")
        if sub_choice == '1':
            calculate_average_over_high()
        elif sub_choice == '2':
            matrix_dict_analysis = load_matrix_dict_analysis()
            all_group_names = list(matrix_dict_analysis.keys())
            if not all_group_names:
                print("沒有任何族群資料可供分析。")
                continue
            print("開始分析所有族群中的股票...")
            all_group_over_high_averages = []

            for i, group in enumerate(all_group_names):
                print(f"\n=== 分析族群：{group} ===")
                group_average = calculate_average_over_high(group_name=group)
                if group_average is not None:
                    all_group_over_high_averages.append(group_average)
                    
            if all_group_over_high_averages:
                overall_group_average = sum(all_group_over_high_averages) / len(all_group_over_high_averages)
                print(f"\n全部族群的平均過高間隔：{overall_group_average:.2f} 分鐘")
            else:
                print("\n沒有任何族群發生過高間隔的情形。")
        elif sub_choice == '0':
            main_menu()
        else:
            print("無效的選項，請重新輸入")

def load_kline_data():
    daily_kline_data = {}
    intraday_kline_data = {}

    if os.path.exists('daily_kline_data.json'):
        with open('daily_kline_data.json', 'r', encoding='utf-8') as f:
            try:
                daily_kline_data = json.load(f)
                if not daily_kline_data:
                    print("日K線數據檔案為空，請先更新數據。")
            except json.JSONDecodeError:
                print("日K線數據檔案格式錯誤，請先更新數據。")

    if os.path.exists('intraday_kline_data.json'):
        with open('intraday_kline_data.json', 'r', encoding='utf-8') as f:
            try:
                intraday_kline_data = json.load(f)
                if not intraday_kline_data:
                    print("一分K線數據檔案為空，請先更新數據。")
            except json.JSONDecodeError:
                print("一分K線數據檔案格式錯誤，請先更新數據。")

    return daily_kline_data, intraday_kline_data

def calculate_average_over_high(group_name=None):
    daily_kline_data, intraday_kline_data = load_kline_data()

    matrix_dict_analysis = load_matrix_dict_analysis()
    
    if group_name is None:
        group_name = input("請輸入要分析的族群名稱：")
    
    if group_name not in matrix_dict_analysis:
        print("沒有此族群資料")
        return None

    symbols_to_analyze = matrix_dict_analysis[group_name]
    disposition_stocks = load_disposition_stocks()
    symbols_to_analyze = [symbol for symbol in symbols_to_analyze if symbol not in disposition_stocks]

    if not symbols_to_analyze:
        print(f"{group_name} 中沒有可供分析的股票。")
        return None

    print(f"開始分析族群 {group_name} 中的股票...")
    any_condition_one_triggered = False 
    group_over_high_averages = []

    for symbol in symbols_to_analyze:
        print(f"\n正在分析股票：{symbol}")
        
        if symbol not in daily_kline_data or symbol not in intraday_kline_data:
            print(f"無法取得 {symbol} 的日 K 線或一分 K 線數據，跳過。")
            continue
        
        daily_kline_df = pd.DataFrame(daily_kline_data[symbol])
        intraday_data = pd.DataFrame(intraday_kline_data[symbol])

        condition_one_triggered = False
        condition_two_triggered = False
        previous_high = None
        condition_two_time = None
        over_high_intervals = []

        for idx, row in intraday_data.iterrows():
            current_time = pd.to_datetime(row['time']).time()
            if previous_high is None:
                previous_high = row['high']
                continue

            if not condition_one_triggered:
                if row['5min_pct_increase'] >= 1.5:
                    condition_one_triggered = True
                    condition_two_triggered = False
                    any_condition_one_triggered = True

                    print(f"{symbol} 觸發條件一，開始監測五分鐘漲幅，五分鐘漲幅: {row['5min_pct_increase']:.2f}%")

            if condition_one_triggered and not condition_two_triggered:
                if row['high'] <= previous_high:
                    current_time_str = current_time.strftime('%H:%M:%S')
                    print(f"{symbol} 觸發條件二！時間：{current_time_str}")

                    condition_two_time = current_time
                    condition_two_triggered = True

            elif condition_two_triggered:
                if row['highest'] > previous_high:
                    condition_three_time_str = current_time.strftime('%H:%M:%S')
                    print(f"{symbol} 觸發條件三！時間：{condition_three_time_str}")
                    if condition_two_time:
                        today = datetime.today().date()
                        condition_two_datetime = datetime.combine(today, condition_two_time)
                        condition_three_datetime = datetime.combine(today, current_time)
                        interval = (condition_three_datetime - condition_two_datetime).total_seconds() / 60
                        print(f"{symbol} 過高間隔：{interval:.2f} 分鐘")
                        over_high_intervals.append(interval)

                    condition_one_triggered = False
                    condition_two_triggered = False
                    condition_two_time = None

            previous_high = row['high']

        if over_high_intervals:
            q1 = np.percentile(over_high_intervals, 25)
            q3 = np.percentile(over_high_intervals, 75)
            iqr = q3 - q1
            lower_bound = q1 - 1.5 * iqr
            upper_bound = q3 + 1.5 * iqr
            filtered_intervals = [interval for interval in over_high_intervals if lower_bound <= interval <= upper_bound]
            if filtered_intervals:
                average_interval = sum(filtered_intervals) / len(filtered_intervals)
                print(f"{symbol} 平均過高間隔：{average_interval:.2f} 分鐘")
                group_over_high_averages.append(average_interval)
            else:
                print(f"{symbol} 沒有有效的過高間隔數據")
        else:
            print(f"{symbol} 沒有觸發過高間隔的情形")

    if group_over_high_averages:
        group_average_over_high = sum(group_over_high_averages) / len(group_over_high_averages)
        print(f"{group_name} 平均過高間隔：{group_average_over_high:.2f} 分鐘")
        return group_average_over_high
    else:
        print(f"{group_name} 沒有有效的過高間隔數據")
        return None

def main_menu():
    global capital_per_stock
    load_settings()
    while True:
        print("\n請選擇功能：")
        print("1. 計算平均過高")
        print("2. 自行選擇進場模式")
        print("3. 極大化利潤模式")
        print("4. 管理族群")
        print("5. 查詢處置股清單")
        print("6. 更新K線數據")
        print("7. 設定")
        print("8. 開始交易")
        print("9. 退出程式")
        choice = input("請輸入選項：")
        if choice == '1':
            calculate_average_over_high_list()
        elif choice == '2':
            simulate_trading_menu()
        elif choice == '3':
            maximize_profit_analysis()
        elif choice == '4':
            manage_groups()
        elif choice == '5':
            display_disposition_stocks()
        elif choice == '6':
            update_kline_data_menu()
        elif choice == '7':
            settings_menu()
        elif choice == '8':
            start_trading()
        elif choice == '9':
            print("退出程式")
            break
        else:
            print("無效的選項，請重新輸入")

capital_per_stock = 0
transaction_fee = 0
transaction_discount = 0
trading_tax = 0
below_50 = 0
price_gap_50_to_100 = 0
price_gap_100_to_500 = 0
price_gap_500_to_1000 = 0
price_gap_above_1000 = 0
allow_reentry_after_stop_loss = False

def load_symbols_to_analyze():
    nb_matrix_dict = load_nb_matrix_dict()
    consolidated_symbols = nb_matrix_dict.get("consolidated_symbols", {})
    symbols = []
    for group_symbols in consolidated_symbols.values():
        symbols.extend(group_symbols)
    disposition_stocks = load_disposition_stocks()
    symbols = [symbol for symbol in symbols if symbol not in disposition_stocks]
    return symbols

def start_trading():
    client, api_key = init_fugle_client()
    symbols_to_analyze = load_symbols_to_analyze()
    stop_trading = False
    accumulated_messages = []
    max_symbols_to_fetch = 20  # 設定要取得的股票數量

    # 載入現有的 auto_daily.json 和 auto_intraday.json
    existing_auto_daily_data = {}
    if os.path.exists('auto_daily.json'):
        with open('auto_daily.json', 'r', encoding='utf-8') as f:
            try:
                existing_auto_daily_data = json.load(f)
            except json.JSONDecodeError:
                existing_auto_daily_data = {}

    auto_daily_data = {}
    data_is_same = True
    count = 0
    symbols_fetched = 0

    # 初始數據獲取與比對
    print("開始取得日K線數據並與現有資料比對...")
    for symbol in symbols_to_analyze:
        if symbols_fetched >= max_symbols_to_fetch:
            break  # 只取得20支股票的數據
        if count >= 55:
            print("已達到55次API請求，休息1分鐘...")
            time_module.sleep(60)
            count = 0
        start_date = (datetime.now() - timedelta(days=30)).strftime('%Y-%m-%d')  # 確保有足夠的數據
        end_date = datetime.now().strftime('%Y-%m-%d')
        print(f"正在取得 {symbol} 從 {start_date} 到 {end_date} 的日K數據...")
        daily_kline_df = fetch_daily_kline_data(client, symbol, days=20)
        count += 1
        if daily_kline_df.empty:
            print(f"無法取得 {symbol} 的日K數據，跳過。")
            continue
        daily_kline_data = daily_kline_df.to_dict(orient='records')
        auto_daily_data[symbol] = daily_kline_data
        existing_data = existing_auto_daily_data.get(symbol)
        if existing_data != daily_kline_data:
            data_is_same = False
            print(f"{symbol} 的日K數據與現有資料不同，將更新資料。")
        else:
            print(f"{symbol} 的日K數據與現有資料相同，跳過更新。")
        symbols_fetched += 1

    if symbols_fetched < max_symbols_to_fetch:
        print(f"注意：僅取得了 {symbols_fetched} 支股票的日K數據。")

    if data_is_same:
        print("所有20支股票的日K數據與現有資料相同，跳過獲取資料，開始監控。")
        auto_intraday_data = {}
        if os.path.exists('auto_intraday.json'):
            with open('auto_intraday.json', 'r', encoding='utf-8') as f:
                try:
                    auto_intraday_data = json.load(f)
                except json.JSONDecodeError:
                    auto_intraday_data = {}
    else:
        print("有股票的日K數據更新，更新 auto_daily.json 和 auto_intraday.json。")
        with open('auto_daily.json', 'w', encoding='utf-8') as f:
            json.dump(auto_daily_data, f, ensure_ascii=False, indent=4, default=str)

        # 更新 auto_intraday.json
        print("開始取得一分K數據並更新 auto_intraday.json...")
        auto_intraday_data = {}
        symbols_fetched_intraday = 0
        for symbol in symbols_to_analyze[:max_symbols_to_fetch]:
            if symbols_fetched_intraday >= max_symbols_to_fetch:
                break
            if count >= 55:
                print("已達到55次API請求，休息1分鐘...")
                time_module.sleep(60)
                count = 0
            print(f"正在取得 {symbol} 的一分K數據...")
            intraday_df = fetch_intraday_data(client, symbol, get_recent_trading_day(), daily_kline_df)
            count += 1
            if intraday_df.empty:
                print(f"無法取得 {symbol} 的一分K數據，跳過。")
                continue
            auto_intraday_data[symbol] = intraday_df.to_dict(orient='records')
            print(f"{symbol} 的一分K數據已成功處理並返回。")
            print(f"已取得 {symbol} 的一分K數據並加入 auto_intraday.json")
            symbols_fetched_intraday += 1

        with open('auto_intraday.json', 'w', encoding='utf-8') as f:
            json.dump(auto_intraday_data, f, ensure_ascii=False, indent=4, default=str)
        print("已更新 auto_daily.json 和 auto_intraday.json")

    # 開始監控
    print("開始監控，即時取得一分K數據。")

    while not stop_trading:
        current_time = datetime.now()
        current_weekday = current_time.weekday()
        market_open_time = datetime.strptime("09:00", "%H:%M").time()
        market_close_time = datetime.strptime("13:30", "%H:%M").time()
        is_weekday = current_weekday < 5
        in_market_hours = market_open_time <= current_time.time() <= market_close_time

        if is_weekday and in_market_hours:
            # 獲取即時數據的程式碼
            message = f"{current_time.strftime('%Y-%m-%d %H:%M:%S')} 市場開盤中，取得即時一分K數據。"
            accumulated_messages.append(message)
            count = 0
            for symbol in symbols_to_analyze[:max_symbols_to_fetch]:
                if count >= 55:
                    print("已達到55次API請求，休息1分鐘...")
                    time_module.sleep(60)
                    count = 0
                latest_candle = fetch_latest_intraday_data(client, symbol)
                count += 1
                if latest_candle:
                    if symbol not in auto_intraday_data:
                        auto_intraday_data[symbol] = []
                    auto_intraday_data[symbol].append(latest_candle)
                    print(f"已取得 {symbol} 的一分K數據並加入 auto_intraday.json")
            with open('auto_intraday.json', 'w', encoding='utf-8') as f:
                json.dump(auto_intraday_data, f, ensure_ascii=False, indent=4, default=str)
        else:
            message = "非開盤時間，沒有K線數據"
            accumulated_messages.append(message)

        # 清除控制台並輸出累積的訊息
        os.system('cls')  # 清除控制台（Windows）
        print('\n'.join(accumulated_messages))
        print("輸入 'Q' 返回主選單：", end='', flush=True)

        # 非阻塞式輸入檢測
        start_time = time_module.time()
        while time_module.time() - start_time < 60:
            if msvcrt.kbhit():
                key = msvcrt.getwch()
                if key.upper() == 'Q':
                    print("\n收到退出指令，停止交易...")
                    stop_trading = True
                    break
            time_module.sleep(1)
    print("已停止交易，返回主選單")


def fetch_latest_intraday_data(client, symbol):
    try:
        candles_response = client.stock.intraday.candles(
            symbol=symbol,
            oddLot=False,
            timeframe='1',
            limit=1
        )
        if candles_response and candles_response.get('data'):
            candle = candles_response['data'][0]
            candle['symbol'] = symbol
            candle['date'] = datetime.now().date().isoformat()
            candle['time'] = datetime.now().strftime('%H:%M:%S')
            return candle
        else:
            print(f"無法取得 {symbol} 的最新一分K數據，API回應中無資料")
            return None
    except Exception as e:
        print(f"獲取 {symbol} 的最新一分K數據時出錯：{e}")
        return None

def fill_missing_intraday_data(intraday_df, start_time, end_time):
    full_time_index = pd.date_range(start='09:00', end='13:30', freq='1min').time
    if start_time and end_time:
        full_time_index = [t for t in full_time_index if start_time <= t <= end_time]
    
    full_index = pd.MultiIndex.from_product([intraday_df['date'].unique(), full_time_index], names=['date', 'time'])
    intraday_df.set_index(['date', 'time'], inplace=True)
    intraday_df = intraday_df.reindex(full_index)
    
    intraday_df[['symbol', '昨日收盤價', '漲停價']] = intraday_df[['symbol', '昨日收盤價', '漲停價']].ffill().bfill()
    intraday_df['close'] = intraday_df['close'].ffill()
    intraday_df['close'] = intraday_df['close'].fillna(intraday_df['昨日收盤價'])
    intraday_df['open'] = intraday_df['open'].ffill()
    intraday_df['open'] = intraday_df['open'].fillna(intraday_df['close'])
    intraday_df['high'] = intraday_df['high'].ffill()
    intraday_df['high'] = intraday_df['high'].fillna(intraday_df['close'])
    intraday_df['low'] = intraday_df['low'].ffill()
    intraday_df['low'] = intraday_df['low'].fillna(intraday_df['close'])
    intraday_df['volume'] = intraday_df['volume'].fillna(0)
    
    if '5min_pct_increase' not in intraday_df.columns:
        intraday_df['5min_pct_increase'] = 0.0
    else:
        intraday_df['5min_pct_increase'] = intraday_df['5min_pct_increase'].fillna(0.0)
    
    intraday_df.reset_index(inplace=True)
    
    intraday_df['rise'] = (intraday_df['close'] - intraday_df['昨日收盤價']) / intraday_df['昨日收盤價'] * 100
    intraday_df = calculate_5min_pct_increase_and_highest(intraday_df)
    intraday_df['highest'] = intraday_df['highest'].ffill().bfill()
    
    return intraday_df

def fill_historical_intraday_data(client, symbols_to_analyze, start_time=None, end_time=None, fill_previous_day=False):
    auto_intraday_data = {}
    if fill_previous_day:
        trading_day = (datetime.now() - timedelta(days=1)).strftime('%Y-%m-%d')
    else:
        trading_day = get_recent_trading_day()
    
    count = 0

    for symbol in symbols_to_analyze:
        if count >= 55:
            print("已達到55次API請求，休息1分鐘...")
            time_module.sleep(60)
            count = 0

        daily_kline_df = fetch_daily_kline_data(client, symbol, days=2)
        count += 1
        if daily_kline_df.empty:
            print(f"無法取得 {symbol} 的日K數據，跳過。")
            continue

        if count >= 55:
            print("已達到55次API請求，休息1分鐘...")
            time_module.sleep(60)
            count = 0

        intraday_df = fetch_intraday_data(client, symbol, trading_day, daily_kline_df)
        count += 1

        if intraday_df.empty:
            print(f"無法取得 {symbol} 的一分K數據，跳過。")
            continue

        if not fill_previous_day:
            intraday_df = fill_missing_intraday_data(intraday_df, start_time, end_time)

        auto_intraday_data[symbol] = intraday_df.to_dict(orient='records')
        print(f"已取得 {symbol} 的一分K數據並加入 auto_intraday.json")

    save_auto_intraday_data(auto_intraday_data)
    print("已更新 auto_intraday.json")


def monitor_realtime_data(api_key, symbols_to_analyze, in_market=True):
    print("開始即時行情監控...")
    
    auto_intraday_path = 'auto_intraday.json'
    auto_daily_path = 'auto_daily.json'
    
    try:
        if os.path.exists(auto_intraday_path):
            with open(auto_intraday_path, 'r', encoding='utf-8') as f:
                auto_intraday_data = json.load(f)
        else:
            auto_intraday_data = {}
    except Exception as e:
        print(f"無法讀取 auto_intraday.json：{e}")
        auto_intraday_data = {}
    
    symbol_to_last_5_closes = {symbol: [] for symbol in symbols_to_analyze}
    
    current_time = datetime.now().time()
    market_open_time = datetime.strptime("09:00", "%H:%M").time()
    market_close_time = datetime.strptime("13:30", "%H:%M").time()
    in_market_hours = market_open_time <= current_time <= market_close_time

    ws_client = WebSocketClient(api_key=api_key)
    stock = ws_client.stock

    terminate_flag = threading.Event()

    def handle_message(message):
        try:
            if message.get('event') == 'data' and message.get('channel') == 'candles':
                candle_data = message['data']
                symbol = candle_data.get('symbol')
                if symbol and symbol in symbols_to_analyze:
                    close_price = candle_data.get('close')
                    if close_price is not None and isinstance(close_price, (int, float)):
                        if symbol not in auto_intraday_data:
                            auto_intraday_data[symbol] = []
                        auto_intraday_data[symbol].append(candle_data)
                        print(f"已更新 {symbol} 的即時K線數據")
                        save_auto_intraday_data(auto_intraday_data)
                    else:
                        print(f"收到的數據缺少 'close' 或 'close' 不是數字，忽略 {symbol} 的數據")
                else:
                    if in_market_hours:
                        print(f"收到的訊息缺少 'symbol' 或不在待分析列表中，忽略該訊息。")
            else:
                pass
        except Exception as e:
            print(f"處理訊息時發生錯誤：{e}")
    
    stock.on('message', handle_message)
    
    ws_client.connect_async()
    ws_client.start()
    
    for symbol in symbols_to_analyze:
        stock.subscribe({
            'channel': 'candles',
            'symbol': symbol
        })
        print(f"已訂閱 {symbol} 的即時K線數據")
    
    def listen_for_quit():
        while not terminate_flag.is_set():
            user_input = input()
            if user_input.strip().upper() == 'Q':
                print("收到退出指令，關閉監控...")
                terminate_flag.set()
                ws_client.stop()
                break
    
    quit_thread = threading.Thread(target=listen_for_quit, daemon=True)
    quit_thread.start()
    
    try:
        while not terminate_flag.is_set():
            time_module.sleep(1)
    except KeyboardInterrupt:
        print("停止即時行情監控")
        terminate_flag.set()
        ws_client.stop()
    print("已退出即時行情監控。")

def handle_message(message):
    print(message)

def initialize_websocket_client(api_token):
    """
    初始化 WebSocket 客戶端
    """
    try:
        ws_client = WebSocketClient(api_key=api_token)
        return ws_client
    except Exception as e:
        print(f"WebSocket 初始化失敗：{e}")
        return None

def fetch_intraday_kline_data(ws_client, symbol):
    """
    從 WebSocket 獲取即時 K 線數據
    """
    try:
        ws_client.subscribe(data_type='candles', symbolId=symbol, timeframe='1')
        latest_candle = ws_client.get_latest_candle(symbol)
        return latest_candle
    except Exception as e:
        print(f"在獲取 {symbol} 即時 K 線數據時出現錯誤：{e}")
        return None
    
def update_auto_daily_data(client, symbols_to_analyze):
    auto_daily_data = {}
    for symbol in symbols_to_analyze:
        daily_kline_df = fetch_daily_kline_data(client, symbol, days=2)
        if not daily_kline_df.empty:
            auto_daily_data[symbol] = daily_kline_df.to_dict(orient='records')
        else:
            print(f"無法取得 {symbol} 的日K線資料。")
    with open('auto_daily.json', 'w', encoding='utf-8') as f:
        json.dump(auto_daily_data, f, ensure_ascii=False, indent=4)
    return auto_daily_data

def calculate_5min_pct_increase(auto_intraday_data, symbol):
    data = auto_intraday_data[symbol]
    num_data = len(data)
    
    if num_data == 1:
        data[-1]['5min_pct_increase'] = 0
    else:
        if num_data < 5:
            subset = data
        else:
            subset = data[-5:]
        
        required_fields = ['close', 'high', 'low']
        if not all(field in subset[0] for field in required_fields):
            data[-1]['5min_pct_increase'] = None
            return
        
        start_price = subset[0]['close']
        end_price = subset[-1]['close']
        highest_price = max(item['high'] for item in subset if 'high' in item and item['high'] is not None)
        lowest_price = min(item['low'] for item in subset if 'low' in item and item['low'] is not None)
        
        if end_price >= start_price:
            if lowest_price != 0:
                pct_increase = ((highest_price - lowest_price) / lowest_price) * 100
            else:
                pct_increase = 0
        else:
            if highest_price != 0:
                pct_increase = -((highest_price - lowest_price) / highest_price) * 100
            else:
                pct_increase = 0
        
        data[-1]['5min_pct_increase'] = pct_increase

def save_auto_intraday_data(auto_intraday_data):
    try:
        with open('auto_intraday.json', 'w', encoding='utf-8') as f:
            json.dump(auto_intraday_data, f, ensure_ascii=False, indent=4, default=str)
        print("已成功儲存 auto_intraday.json")
    except Exception as e:
        print(f"儲存 auto_intraday.json 時發生錯誤：{e}")


def update_kline_data_menu():
    while True:
        print("\n更新K線數據選單：")
        print("1. 更新K線數據")
        print("2. 查看K線數據")
        print("0. 返回主選單")
        choice = input("請輸入選項：")
        if choice == '1':
            update_kline_data()
        elif choice == '2':
            view_kline_data()
        elif choice == '0':
            main_menu()
        else:
            print("無效的選項，請重新輸入")

def convert_datetime_to_str(obj):
    if isinstance(obj, dict):
        return {k: convert_datetime_to_str(v) for k, v in obj.items()}
    elif isinstance(obj, list):
        return [convert_datetime_to_str(element) for element in obj]
    elif isinstance(obj, (datetime, pd.Timestamp, time, date)):
        return obj.isoformat()
    else:
        return obj

def update_kline_data():
    client, api_key = init_fugle_client()
    matrix_dict_analysis = load_matrix_dict_analysis()
    if not matrix_dict_analysis:
        print("沒有任何族群資料，請先管理族群。")
        return

    print("正在更新處置股清單...")
    fetch_disposition_stocks(client, matrix_dict_analysis)
    print("處置股清單已更新。")

    disposition_stocks = load_disposition_stocks()

    all_symbols = []
    for group, symbols in matrix_dict_analysis.items():
        filtered_symbols = [symbol for symbol in symbols if symbol not in disposition_stocks]
        all_symbols.extend(filtered_symbols)

    if not all_symbols:
        print("過濾後沒有任何可供分析的股票。")
        return

    daily_kline_data = {}
    intraday_kline_data = {}
    count = 0

    for symbol in all_symbols:
        if count >= 55:
            print("已達到55次API請求，休息1分鐘...")
            time_module.sleep(60)
            count = 0

        daily_kline_df = fetch_daily_kline_data(client, symbol, days=2)
        count += 1

        if daily_kline_df.empty:
            print(f"無法取得 {symbol} 的日K數據，跳過。")
            continue
        daily_kline_data[symbol] = daily_kline_df.to_dict(orient='records')

        if count >= 55:
            print("已達到55次API請求，休息1分鐘...")
            time_module.sleep(60)
            count = 0

        recent_day = get_recent_trading_day()
        intraday_df = fetch_intraday_data(client, symbol, recent_day, daily_kline_df)
        count += 1

        if intraday_df.empty:
            print(f"無法取得 {symbol} 的一分K數據，跳過。")
            continue
        intraday_kline_data[symbol] = intraday_df.to_dict(orient='records')
        print(f"已取得 {symbol} 的一分K數據並加入 intraday_kline_data.json")

    daily_kline_data_str = convert_datetime_to_str(daily_kline_data)
    intraday_kline_data_str = convert_datetime_to_str(intraday_kline_data)

    with open('daily_kline_data.json', 'w', encoding='utf-8') as f:
        json.dump(daily_kline_data_str, f, indent=4, ensure_ascii=False, default=str)
    with open('intraday_kline_data.json', 'w', encoding='utf-8') as f:
        json.dump(intraday_kline_data_str, f, indent=4, ensure_ascii=False, default=str)

    print("K線數據已儲存。")

    print("正在計算每個族群的相似度...")
    similarity_df = calculate_kline_similarity([pd.DataFrame(data) for data in intraday_kline_data.values()])
    mt_matrix_dict = similarity_df.to_dict(orient='records')
    save_mt_matrix_dict(mt_matrix_dict)
    print("相似度計算完成並已儲存至 mt_matrix_dict.json。")

    consolidate_and_save_stock_symbols()
    print("股票代號已統整並儲存至 nb_matrix_dict.json，按族群分類。")

    print("K線數據更新完成。")


def view_kline_data():
    if not os.path.exists('intraday_kline_data.json'):
        print("尚未更新一分K數據，請先更新K線數據。")
        return
    with open('intraday_kline_data.json', 'r', encoding='utf-8') as f:
        intraday_kline_data = json.load(f)
    
    for symbol, data in intraday_kline_data.items():
        print(f"\n股票代號：{symbol} 的一分K數據：")
        df = pd.DataFrame(data)
        if df.empty:
            print("沒有資料。")
            continue
        
        if 'time' in df.columns:
            try:
                df['time'] = pd.to_datetime(df['time'])
            except Exception as e:
                print(f"轉換時間欄位時發生錯誤：{e}")
                continue
        
        print(df)

def set_price_gap_stop_loss():
    global price_gap_below_50, price_gap_50_to_100, price_gap_100_to_500, price_gap_500_to_1000, price_gap_above_1000
    price_gap_below_50 = int(input(f"50元以下股票停損價差，目前為 {price_gap_below_50} 元："))
    price_gap_50_to_100 = int(input(f"50~100元股票停損價差，目前為 {price_gap_50_to_100} 元："))
    price_gap_100_to_500 = int(input(f"100~500元股票停損價差，目前為 {price_gap_100_to_500} 元："))
    price_gap_500_to_1000 = int(input(f"500~1000元股票停損價差，目前為 {price_gap_500_to_1000} 元："))
    price_gap_above_1000 = int(input(f"1000元以上股票停損價差，目前為 {price_gap_above_1000} 元："))
    save_settings()

def save_settings():
    with open('settings.json', 'w', encoding='utf-8') as f:
        json.dump({
            'capital_per_stock': capital_per_stock,
            'transaction_fee': transaction_fee,
            'transaction_discount': transaction_discount,
            'trading_tax': trading_tax,
            'below_50': below_50,
            'price_gap_50_to_100': price_gap_50_to_100,
            'price_gap_100_to_500': price_gap_100_to_500,
            'price_gap_500_to_1000': price_gap_500_to_1000,
            'price_gap_above_1000': price_gap_above_1000,
            'allow_reentry_after_stop_loss': allow_reentry_after_stop_loss
        }, f, indent=4)

def load_settings():
    global capital_per_stock, transaction_fee, transaction_discount, trading_tax
    global below_50, price_gap_50_to_100, price_gap_100_to_500, price_gap_500_to_1000, price_gap_above_1000
    global allow_reentry_after_stop_loss
    if os.path.exists('settings.json'):
        with open('settings.json', 'r', encoding='utf-8') as f:
            settings = json.load(f)
            capital_per_stock = settings.get('capital_per_stock', 0)
            transaction_fee = settings.get('transaction_fee', 0)
            transaction_discount = settings.get('transaction_discount', 0)
            trading_tax = settings.get('trading_tax', 0)
            below_50 = settings.get('below_50', 0)
            price_gap_50_to_100 = settings.get('price_gap_50_to_100', 0)
            price_gap_100_to_500 = settings.get('price_gap_100_to_500', 0)
            price_gap_500_to_1000 = settings.get('price_gap_500_to_1000', 0)
            price_gap_above_1000 = settings.get('price_gap_above_1000', 0)
            allow_reentry_after_stop_loss = settings.get('allow_reentry_after_stop_loss', False)
    else:
        capital_per_stock = 1000
        transaction_fee = 0.1425
        transaction_discount = 20.0
        trading_tax = 0.15
        below_50 = 500
        price_gap_50_to_100 = 1000
        price_gap_100_to_500 = 2000
        price_gap_500_to_1000 = 3000
        price_gap_above_1000 = 5000
        allow_reentry_after_stop_loss = False

def settings_menu():
    global capital_per_stock, transaction_fee, transaction_discount, trading_tax
    global below_50, price_gap_50_to_100, price_gap_100_to_500, price_gap_500_to_1000, price_gap_above_1000
    global allow_reentry_after_stop_loss
    while True:
        print("\n設定選單：")
        print(f"1. 設定每檔股票投入資本額（目前為 {capital_per_stock} 萬元）")
        print(f"2. 手續費設定，目前為 {transaction_fee}%")
        print(f"3. 手續費折數設定，目前為 {transaction_discount}%")
        print(f"4. 證交稅設定，目前為 {trading_tax}%")
        print("5. 價差停損設定")
        print("6. 停損再進場設定")
        print("0. 返回主選單")
        choice = input("請輸入選項：")
        if choice == "1":
            set_capital_per_stock()
        elif choice == "2":
            transaction_fee = float(input("請輸入手續費（%）："))
            save_settings()
        elif choice == "3":
            transaction_discount = float(input("請輸入手續費折數（%）："))
            save_settings()
        elif choice == "4":
            trading_tax = float(input("請輸入證交稅（%）："))
            save_settings()
        elif choice == "5":
            price_gap_stop_loss_menu()
        elif choice == "6":
            stop_loss_reentry_menu()
        elif choice == "0":
            main_menu()
        else:
            print("無效的選項，請重新輸入")

def stop_loss_reentry_menu():
    global allow_reentry_after_stop_loss
    while True:
        status = "開啟" if allow_reentry_after_stop_loss else "關閉"
        print(f"\n目前為({status}停損後進場)")
        print("1.開啟停損後進場")
        print("2.關閉停損後進場")
        print("3.返回上一頁")
        choice = input("請輸入選項：")
        if choice == '1':
            allow_reentry_after_stop_loss = True
            print("已開啟停損後進場功能")
            save_settings()
        elif choice == '2':
            allow_reentry_after_stop_loss = False
            print("已關閉停損後進場功能")
            save_settings()
        elif choice == '3':
            settings_menu()
        else:
            print("無效的選項，請重新輸入")

def price_gap_stop_loss_menu():
    global below_50, price_gap_50_to_100, price_gap_100_to_500, price_gap_500_to_1000, price_gap_above_1000
    while True:
        print(f"1. 50元以下股票停損價差，目前為 {below_50} 元")
        print(f"2. 50~100元股票停損價差，目前為 {price_gap_50_to_100} 元")
        print(f"3. 100~500元股票停損價差，目前為 {price_gap_100_to_500} 元")
        print(f"4. 500~1000元股票停損價差，目前為 {price_gap_500_to_1000} 元")
        print(f"5. 1000元以上股票停損價差，目前為 {price_gap_above_1000} 元")
        print("6. 返回上一頁")
        choice = input("請選擇要設定的項目：")
        if choice == "1":
            below_50 = float(input("請輸入50元以下股票的停損價差："))
        elif choice == "2":
            price_gap_50_to_100 = float(input("請輸入50~100元股票的停損價差："))
        elif choice == "3":
            price_gap_100_to_500 = float(input("請輸入100~500元股票的停損價差："))
        elif choice == "4":
            price_gap_500_to_1000 = float(input("請輸入500~1000元股票的停損價差："))
        elif choice == "5":
            price_gap_above_1000 = float(input("請輸入1000元以上股票的停損價差："))
        elif choice == "6":
            break
        else:
            print("無效選擇，請重試。")
        save_settings()

def load_capital_per_stock():
    if os.path.exists('capital_per_stock.json'):
        with open('capital_per_stock.json', 'r', encoding='utf-8') as f:
            return json.load(f).get('capital_per_stock', 0)
    else:
        return 0
    
def save_capital_per_stock():
    global capital_per_stock
    with open('capital_per_stock.json', 'w', encoding='utf-8') as f:
        json.dump({'capital_per_stock': capital_per_stock}, f)

def simulate_trading_menu():
    matrix_dict_analysis = load_matrix_dict_analysis()
    if not matrix_dict_analysis:
        print("沒有族群資料，請先管理族群。")
        return

    while True:
        print("請選擇操作：")
        print("1. 分析單一族群")
        print("2. 分析全部族群")
        print("0. 返回主選單")
        choice = input("請輸入選項編號：")

        if choice == '1':
            group_name = input("請輸入要分析的族群名稱：")
            if group_name not in matrix_dict_analysis:
                print("沒有此族群資料")
                continue

            try:
                wait_minutes = int(input("請輸入等待時間（分鐘）："))
            except ValueError:
                print("等待時間必須是整數。")
                continue

            hold_minutes_input = input("請輸入持有時間（分鐘，輸入 'F' 代表持有到13:30強制出場）：")
            if hold_minutes_input.upper() == 'F':
                hold_minutes = None
            else:
                try:
                    hold_minutes = int(hold_minutes_input)
                except ValueError:
                    print("持有時間必須是整數或 'F'。")
                    continue

            disposition_stocks = load_disposition_stocks()
            symbols_to_analyze = matrix_dict_analysis[group_name]
            symbols_to_analyze = [symbol for symbol in symbols_to_analyze if symbol not in disposition_stocks]
            if len(symbols_to_analyze) == 0:
                print(f"{group_name} 中沒有可供分析的股票。")
                continue

            daily_kline_data, intraday_kline_data = load_kline_data()

            stock_data_collection = initialize_stock_data(symbols_to_analyze, daily_kline_data, intraday_kline_data)
            if not stock_data_collection:
                print("無法獲取有效的一分 K 資料，無法進行分析")
                continue

            total_profit, avg_profit_rate = process_group_data(stock_data_collection, wait_minutes, hold_minutes, matrix_dict_analysis, verbose=True)

            print(f"\n模擬交易完成，總利潤：{int(total_profit) if total_profit is not None else 0} 元，平均報酬率：{avg_profit_rate if avg_profit_rate is not None else 0:.2f}%\n")

        elif choice == '2':
            try:
                wait_minutes = int(input("請輸入等待時間（分鐘）："))
            except ValueError:
                print("等待時間必須是整數。")
                continue

            hold_minutes_input = input("請輸入持有時間（分鐘，輸入 'F' 代表持有到13:30強制出場）：")
            if hold_minutes_input.upper() == 'F':
                hold_minutes = None
            else:
                try:
                    hold_minutes = int(hold_minutes_input)
                except ValueError:
                    print("持有時間必須是整數或 'F'。")
                    continue

            day_total_profit = 0
            day_avg_profit_rates = []

            for group_name in matrix_dict_analysis.keys():
                print(f"\n正在分析族群：{group_name}")

                disposition_stocks = load_disposition_stocks()
                symbols_to_analyze = matrix_dict_analysis[group_name]
                symbols_to_analyze = [symbol for symbol in symbols_to_analyze if symbol not in disposition_stocks]
                if len(symbols_to_analyze) == 0:
                    print(f"{group_name} 中沒有可供分析的股票。")
                    continue

                daily_kline_data, intraday_kline_data = load_kline_data()

                stock_data_collection = initialize_stock_data(symbols_to_analyze, daily_kline_data, intraday_kline_data)
                if not stock_data_collection:
                    print(f"無法獲取 {group_name} 的有效一分 K 資料，跳過。")
                    continue

                total_profit, avg_profit_rate = process_group_data(stock_data_collection, wait_minutes, hold_minutes, matrix_dict_analysis, verbose=True)

                if total_profit is not None and avg_profit_rate is not None:
                    day_total_profit += total_profit
                    day_avg_profit_rates.append(avg_profit_rate)
                else:
                    pass

                print(f"族群 {group_name} 的模擬交易完成，總利潤：{int(total_profit) if total_profit is not None else 0} 元，平均報酬率：{avg_profit_rate if avg_profit_rate is not None else 0:.2f}%")

            if day_avg_profit_rates:
                day_avg_profit_rate = sum(day_avg_profit_rates) / len(day_avg_profit_rates)
            else:
                day_avg_profit_rate = 0.0
            
            if day_total_profit > 0:
                print(f"{RED}={RESET}" * 50)
                print(f"{RED}\n當日總利潤：{int(day_total_profit)} 元{RESET}")
                print(f"{RED}當日報酬率：{day_avg_profit_rate:.2f}%\n{RESET}")
                print(f"{RED}={RESET}" * 50)
            elif day_total_profit < 0:
                print(f"{GREEN}={RESET}" * 50)
                print(f"{GREEN}\n當日總利潤：{int(day_total_profit)} 元{RESET}")
                print(f"{GREEN}當日報酬率：{day_avg_profit_rate:.2f}%\n{RESET}")
                print(f"{GREEN}={RESET}" * 50)
            else:
                print("=" * 50)
                print(f"\n當日總利潤：{int(day_total_profit)} 元")
                print(f"當日報酬率：{day_avg_profit_rate:.2f}%\n")
                print("=" * 50)

        elif choice == '0':
            break
        else:
            print("無效的選項，請重新輸入。")

def display_disposition_stocks():
    disposition_file = 'Disposition.json'
    try:
        with open(disposition_file, 'r', encoding='utf-8') as f:
            disposition_data = json.load(f)
            if isinstance(disposition_data, list):
                stock_codes = disposition_data
            elif isinstance(disposition_data, dict):
                stock_codes = disposition_data.get("stock_codes", [])
            else:
                print(f"錯誤：{disposition_file} 文件格式不正確。")
                return
    except FileNotFoundError:
        print(f"錯誤：無法找到 {disposition_file} 文件。")
        return
    except json.JSONDecodeError:
        print(f"錯誤：{disposition_file} 文件格式不正確。")
        return

    if not stock_codes:
        print(f"{disposition_file} 中沒有任何股票代號。")
        return

    items_per_page = 10
    total_items = len(stock_codes)
    total_pages = (total_items + items_per_page - 1) // items_per_page
    current_page = 1

    while True:
        start_idx = (current_page - 1) * items_per_page
        end_idx = start_idx + items_per_page
        page_items = stock_codes[start_idx:end_idx]

        print("\n" + "=" * 50)
        print(f"{disposition_file} 股票代號列表 - 第 {current_page} 頁 / 共 {total_pages} 頁")
        print("=" * 50)
        for idx, code in enumerate(page_items, start=1 + start_idx):
            print(f"{idx}. {code}")
        print("=" * 50)
        if total_pages == 1:
            print("已顯示所有股票代號。")
            break

        print("導航選項：")
        if current_page > 1:
            print("P - 上一頁")
        if current_page < total_pages:
            print("N - 下一頁")
        print("0 - 返回主選單")

        choice = input("請輸入選項（N/P/0）：").strip().upper()

        if choice == 'N' and current_page < total_pages:
            current_page += 1
        elif choice == 'P' and current_page > 1:
            current_page -= 1
        elif choice == '0':
            break
        else:
            print("無效的選項，請重新輸入。")

def set_capital_per_stock():
    global capital_per_stock
    capital_per_stock = int(input("請輸入每檔投入資本額（萬元）："))
    print(f"每檔投入資本額已設定為：{capital_per_stock} 萬元")
    save_settings()

def maximize_profit_analysis():
    print("進入極大化利潤模式...")
    
    matrix_dict_analysis = load_matrix_dict_analysis()
    if not matrix_dict_analysis:
        print("沒有族群資料，請先管理族群。")
        return

    group_name = input("請輸入要分析的族群名稱：")
    
    if group_name not in matrix_dict_analysis:
        print("沒有此族群資料")
        return

    wait_minutes_range = range(1, 10)
    hold_minutes_range = list(range(60, 120)) + ['F']

    disposition_stocks = load_disposition_stocks()
    symbols_to_analyze = matrix_dict_analysis[group_name]
    symbols_to_analyze = [symbol for symbol in symbols_to_analyze if symbol not in disposition_stocks]
    if len(symbols_to_analyze) == 0:
        print(f"{group_name} 中沒有可供分析的股票。")
        return

    daily_kline_data, intraday_kline_data = load_kline_data()

    stock_data_collection = initialize_stock_data(symbols_to_analyze, daily_kline_data, intraday_kline_data)
    if not stock_data_collection:
        print("無法獲取有效的一分 K 資料，無法進行分析")
        return

    results_df = pd.DataFrame(columns=['等待時間', '持有時間', '總利潤', '平均報酬率'])
    results_df = results_df.astype({
        '等待時間': 'int',
        '持有時間': 'object',
        '總利潤': 'float',
        '平均報酬率': 'float'
    })

    for wait_minutes in wait_minutes_range:
        for hold_minutes in hold_minutes_range:
            print(f"正在分析：等待時間 {wait_minutes} 分鐘、持有時間 {hold_minutes} 分鐘")
            
            total_profit, avg_profit_rate = process_group_data(
                stock_data_collection, wait_minutes, None if hold_minutes == 'F' else hold_minutes, matrix_dict_analysis, verbose=False)
            
            if total_profit is None:
                total_profit = 0.0
            if avg_profit_rate is None:
                avg_profit_rate = 0.0
            
            new_row = pd.DataFrame([{
                '等待時間': wait_minutes,
                '持有時間': hold_minutes,
                '總利潤': float(total_profit),
                '平均報酬率': float(avg_profit_rate)
            }])
            results_df = pd.concat([results_df, new_row], ignore_index=True)

    if results_df.empty:
        print("模擬結果為空，無法進行後續分析。")
        return

    max_profit = results_df['總利潤'].max()
    min_profit = results_df['總利潤'].min()
    best_combination = results_df.loc[results_df['總利潤'].idxmax()]

    print("\n利潤最大的組合：")
    print(f"等待時間：{best_combination['等待時間']} 分鐘，持有時間：{best_combination['持有時間']} 分鐘，總利潤：{int(best_combination['總利潤'])} 元，平均報酬率：{best_combination['平均報酬率']:.2f}%\n")

    pivot_df = results_df.pivot(index='等待時間', columns='持有時間', values='總利潤')

    formatted_pivot_df = pivot_df.copy()
    for col in formatted_pivot_df.columns:
        if col != '等待時間':
            formatted_pivot_df[col] = formatted_pivot_df[col].apply(lambda x: f"{int(x):,}" if pd.notnull(x) else "")

    formatted_pivot_df_reset = formatted_pivot_df.reset_index()

    print("模擬結果：")
    print(tabulate(formatted_pivot_df_reset, headers='keys', tablefmt='psql', showindex=False))

    try:
        with pd.ExcelWriter('模擬結果.xlsx', engine='openpyxl') as writer:
            pivot_df.to_excel(writer, sheet_name='模擬結果', index=True)
            workbook = writer.book
            worksheet = writer.sheets['模擬結果']
            
            max_profit = pivot_df.max().max()
            min_profit = pivot_df.min().min()

            max_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
            min_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')

            for row in worksheet.iter_rows(min_row=2, min_col=2):
                for cell in row:
                    if cell.value == max_profit:
                        cell.fill = max_fill
                    elif cell.value == min_profit:
                        cell.fill = min_fill
        print("\n模擬結果已成功寫入 '模擬結果.xlsx'。")
    except Exception as e:
        print(f"\n寫入 Excel 時發生錯誤：{e}")

def manage_groups():
    current_page = 0
    page_size = 5
    groups = load_matrix_dict_analysis()
    total_pages = (len(groups) + page_size - 1) // page_size

    def display_page(page):
        start = page * page_size
        end = start + page_size
        print("=" * 50)
        print(f"族群及個股列表 - 第 {page + 1} 頁 / 共 {total_pages} 頁")
        print("=" * 50)
        for idx, (group, stocks) in enumerate(list(groups.items())[start:end], start=1):
            print(f"族群: {group}")
            for stock_idx, stock in enumerate(stocks, start=1):
                print(f"  {str(stock_idx).rjust(2)}. {stock}")
            print("-" * 50)
        print("=" * 50)
        if current_page == total_pages - 1:
            print("已顯示所有族群及個股。")
        print("=" * 50)

    while True:
        display_page(current_page)
        print("\nP：上一頁、Q：下一頁、1：新增族群/個股；、2：刪除族群/個股、0：返回主選單")
        choice = input("請選擇操作: ")

        if choice == "P":
            if current_page > 0:
                current_page -= 1
            else:
                print("已經是第一頁！")
        elif choice == "Q":
            if current_page < total_pages - 1:
                current_page += 1
            else:
                print("已經是最後一頁！")
        elif choice == "1":
            add_group_or_stock(groups)
        elif choice == "2":
            delete_group_or_stock(groups)
        elif choice == "0":
            save_matrix_dict(groups)
            break
        else:
            print("無效選項，請重新選擇。")

def add_group_or_stock(groups):
    print("1：新增族群、2：新增族群中的個股")
    choice = input("請選擇操作: ")

    if choice == "1":
        new_group = input("輸入新族群名稱: ")
        if new_group in groups:
            print(f"族群 {new_group} 已存在。")
        else:
            groups[new_group] = []
            print(f"族群 {new_group} 新增成功。")
    elif choice == "2":
        group_name = input("輸入要新增個股的族群名稱: ")
        if group_name in groups:
            new_stock = input("輸入個股代號: ")
            if new_stock in groups[group_name]:
                print(f"個股 {new_stock} 已存在於 {group_name} 族群中。")
            else:
                groups[group_name].append(new_stock)
                print(f"個股 {new_stock} 已新增至 {group_name} 族群。")
        else:
            print(f"族群 {group_name} 不存在。")
    elif choice == "0":
        manage_groups()

def delete_group_or_stock(groups):
    print("1：刪除族群、2：刪除族群中的個股")
    choice = input("請選擇操作: ")

    if choice == "1":
        group_name = input("輸入要刪除的族群名稱: ")
        if group_name in groups:
            del groups[group_name]
            print(f"族群 {group_name} 已刪除。")
        else:
            print(f"族群 {group_name} 不存在。")
    elif choice == "2":
        group_name = input("輸入要刪除個股的族群名稱: ")
        if group_name in groups:
            stock_name = input("輸入要刪除的個股代號: ")
            if stock_name in groups[group_name]:
                groups[group_name].remove(stock_name)
                print(f"個股 {stock_name} 已從 {group_name} 族群中刪除。")
            else:
                print(f"個股 {stock_name} 不存在於 {group_name} 族群中。")
        else:
            print(f"族群 {group_name} 不存在。")
    elif choice == "0":
        manage_groups()

def main():
    load_settings()
    config = load_config("config.yaml")
    client = RestClient(api_key=config['api_key'])
    matrix_dict_analysis = load_matrix_dict_analysis()
    main_menu()

if __name__ == "__main__":
    check_and_install_packages(required_packages)
    
    print("所有必要套件已安裝，開始執行程式...")
    main()